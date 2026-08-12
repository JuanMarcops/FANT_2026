#!/usr/bin/env python3
"""Build the conference Book of Contents from the submissions spreadsheet.

Produces a static website (index.html / index-<lang>.html) and a PDF per
language in the output directory.

Usage:
    python scripts/build.py
        Use the input path(s) from config.yml (input.xlsx_path, falling
        back to input.csv_path if the xlsx is absent).
    python scripts/build.py --input data/some-other-export.xlsx
        Override the input file path (handy for testing with sample data).
    python scripts/build.py --out site
        Choose the output directory (default: site).

All conference-specific settings live in config.yml, not here — see
README.md for how to point this at a different conference.
"""

import argparse
import csv
import re
import shutil
import sys
import unicodedata
from datetime import date
from pathlib import Path

import yaml
from jinja2 import Environment, FileSystemLoader, select_autoescape

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "templates"


def load_config() -> dict:
    with open(ROOT / "config.yml", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def normalize_header(label: str) -> str:
    if label is None:
        return ""
    normalized = unicodedata.normalize("NFKC", label.strip().lower())
    normalized = (
        normalized.replace("ä", "a")
        .replace("ö", "o")
        .replace("ü", "u")
        .replace("ß", "ss")
        .replace("–", "-")
        .replace("—", "-")
        .replace("’", "'")
        .replace("‘", "'")
    )
    return " ".join(normalized.split())


def normalize_column_spec(value):
    if isinstance(value, list):
        return [str(v) for v in value if v is not None]
    if value:
        return [str(value)]
    return []


def build_header_index(header_row: list[str]) -> dict[str, list[int]]:
    index = {}
    for idx, name in enumerate(header_row):
        key = normalize_header(name)
        index.setdefault(key, []).append(idx)
    return index


def find_header_indices(header_index: dict[str, list[int]], candidates: list[str]) -> list[int]:
    indices: list[int] = []
    for candidate in candidates:
        normalized = normalize_header(candidate)
        indices.extend(header_index.get(normalized, []))
    return indices


def get_cell(row: list[str], indices: list[int]) -> str:
    for idx in indices:
        if idx < len(row):
            value = row[idx].strip()
            if value:
                return value
    return ""


def get_multi_cell(row: list[str], indices: list[int]) -> list[str]:
    """Return all non-empty values from the given column indices (preserves order)."""
    return [row[idx].strip() for idx in indices if idx < len(row) and row[idx].strip()]


_NONE_VALUES = frozenset({"none", "n/a", "no", "-", "–", "keine", "no co-author", "no co-authors"})


def clean_none_values(values: list[str]) -> list[str]:
    return [v for v in values if normalize_header(v) not in _NONE_VALUES]


# ── Case normalisation ──────────────────────────────────────────────────────
# Rules applied only when a string is >65 % uppercase letters:
#   1. Mixed-case tokens (e.g. AAArC, SfM) are assumed intentional → kept.
#   2. All-caps tokens that are common stop words → lowercased (except position 0).
#   3. All-caps tokens ≤ 3 letters (not stop words) → assumed acronym → kept.
#   4. Everything else → capitalized (first letter upper, rest lower).

_ACRONYM_MAX = 3

_STOP_WORDS = frozenset({
    # English
    "a", "an", "and", "as", "at", "but", "by", "for", "from",
    "in", "is", "nor", "of", "on", "or", "the", "to", "up", "via", "with",
    # German (after umlaut-stripping: für→fur, über→uber, etc.)
    "am", "an", "auf", "aus", "bei", "das", "dem", "den", "der", "des",
    "die", "ein", "fur", "im", "mit", "nach", "oder", "uber", "und",
    "von", "vor", "zu", "zur",
})


def _normalize_word(word: str, is_first: bool = False) -> str:
    m = re.match(r'^([^A-Za-z0-9]*)(.+?)([^A-Za-z0-9]*)$', word)
    if not m:
        return word
    pre, core, post = m.group(1), m.group(2), m.group(3)
    letters = re.sub(r'[^A-Za-z]', '', core)
    if not letters or not letters.isupper():          # mixed case → intentional
        return word
    if not is_first and letters.lower() in _STOP_WORDS:  # stop word → lowercase
        return pre + core.lower() + post
    if len(letters) <= _ACRONYM_MAX:                  # short all-caps → acronym
        return word
    return pre + core[0].upper() + core[1:].lower() + post


def normalize_case(text: str) -> str:
    """Convert predominantly ALL-CAPS text to capitalised words; leave normal text untouched."""
    if not text:
        return text
    letters = [c for c in text if c.isalpha()]
    if len(letters) < 4:
        return text
    if sum(1 for c in letters if c.isupper()) / len(letters) <= 0.65:
        return text
    words = text.split()
    return ' '.join(_normalize_word(w, i == 0) for i, w in enumerate(words))


_FORMAT_MAP = {
    "presentation":   "Presentations",
    "prasentation":   "Presentations",  # Präsentation
    "vortrag":        "Presentations",
    "paper":          "Presentations",
    "poster":         "Posters",
    "posterbeitrag":  "Posters",
    "poster contribution": "Posters",
    "poster session": "Posters",
    "roundtable":     "Roundtables",
    "round table":    "Roundtables",
    "roundtable discussion": "Roundtables",
}


def normalize_format(value: str) -> str:
    return _FORMAT_MAP.get(normalize_header(value), "")


_COLUMN_DEFAULTS = {
    "authors": ["Author / Presenter", "Author:in/ Vortragende:r"],
    "co_authors": ["Co-authors", "Co-author", "Mitautor:innen"],
    "institution": ["Institutional affiliation", "Institutionelle Zugehörigkeit", "Institution"],
    "title": ["Title of the contribution", "Titel des Beitrags"],
    "abstract": [
        "Abstract (approx. 2–3 sentences)",
        "Abstract (approx. 2-3 sentences)",
        "Abstract (approx. 150 words)",
        "Abstract",
    ],
    "track": [],
    "keywords": ["Keywords (3-5)", "Schlagwörter"],
    "format": ["Format of the contribution", "Format des Beitrags"],
    "language": ["Language of the contribution:", "Sprache des Beitrags:"],
    "first_name": ["First name", "Vorname"],
    "last_name": ["Last name", "Nachname"],
}


def _parse_submission_rows(rows: list[list[str]], colmap: dict) -> list[dict]:
    """Parse an iterable of string rows (header first) into submission dicts."""
    column_candidates = {
        key: normalize_column_spec(colmap.get(key, _COLUMN_DEFAULTS.get(key, []))) or _COLUMN_DEFAULTS.get(key, [])
        for key in _COLUMN_DEFAULTS
    }

    header = rows[0] if rows else []
    header_index = build_header_index(header)

    title_indices = find_header_indices(header_index, column_candidates["title"])
    authors_indices = find_header_indices(header_index, column_candidates["authors"])
    coauthor_indices = find_header_indices(header_index, column_candidates["co_authors"])
    # Google Forms exports overflow co-author inputs as empty-header columns adjacent to a Co-authors column.
    empty_indices = header_index.get("", [])
    for cidx in list(coauthor_indices):
        for eidx in empty_indices:
            if eidx == cidx + 1 and eidx not in coauthor_indices:
                coauthor_indices.append(eidx)
    abstract_indices = find_header_indices(header_index, column_candidates["abstract"])
    track_indices = find_header_indices(header_index, column_candidates["track"])
    keywords_indices = find_header_indices(header_index, column_candidates["keywords"])
    institution_indices = find_header_indices(header_index, column_candidates["institution"])
    format_indices = find_header_indices(header_index, column_candidates["format"])
    first_name_indices = find_header_indices(header_index, column_candidates["first_name"])
    last_name_indices = find_header_indices(header_index, column_candidates["last_name"])

    submissions = []
    for row in rows[1:]:
        if not any(cell.strip() for cell in row):
            continue

        format_value = get_cell(row, format_indices)

        title = normalize_case(get_cell(row, title_indices))
        if not title:
            continue

        abstract = get_cell(row, abstract_indices)
        if not abstract or normalize_header(abstract) in _NONE_VALUES:
            # Skip registrations without a real contribution (e.g. "I just want to attend")
            continue

        authors = normalize_case(get_cell(row, authors_indices))
        if not authors:
            first_name = get_cell(row, first_name_indices)
            last_name = get_cell(row, last_name_indices)
            if first_name or last_name:
                authors = normalize_case(
                    " ".join(part for part in (first_name, last_name) if part)
                )

        co_authors = [normalize_case(v) for v in clean_none_values(get_multi_cell(row, coauthor_indices))]
        institution = normalize_case(get_cell(row, institution_indices))
        keywords = get_cell(row, keywords_indices)
        if normalize_header(keywords) in _NONE_VALUES:
            keywords = ""
        track = get_cell(row, track_indices) or normalize_format(format_value) or "Other"

        submissions.append(
            {
                "authors": authors,
                "co_authors": co_authors,
                "institution": institution,
                "title": title,
                "abstract": abstract,
                "track": track,
                "keywords": keywords,
            }
        )
    return submissions


def read_xlsx(path: Path, colmap: dict) -> list[dict]:
    """Read submissions directly from an xlsx file (first sheet)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [
        [str(cell) if cell is not None else "" for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]
    return _parse_submission_rows(rows, colmap)


def read_csv(path: Path, colmap: dict) -> list[dict]:
    """Read the exported CSV into canonical submission dicts."""
    with open(path, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    return _parse_submission_rows(rows, colmap)


def group_by_session(submissions: list[dict], cfg: dict) -> list[dict]:
    """Return ordered list of {"track": str, "entries": [...]} blocks."""
    order = cfg.get("session_order") or []
    by_track: dict[str, list[dict]] = {}
    for s in submissions:
        by_track.setdefault(s["track"], []).append(s)

    # Sort entries within a track: by last name, or by title when no author.
    for entries in by_track.values():
        entries.sort(key=lambda e: (e["authors"].split() or [e["title"]])[-1].lower())

    ordered = [t for t in order if t in by_track]
    rest = sorted(t for t in by_track if t not in order)
    return [{"track": t, "entries": by_track[t]} for t in ordered + rest]


_PRESENTER_SHEET_NAMES = ["Presenters", "Presentors"]
_SCHEDULE_SHEET_NAMES = ["Schedule"]


def load_schedule(cfg: dict) -> list[dict]:
    """Read the schedule xlsx (input.schedule_xlsx_path) and return rows with presenter info.

    Both sheets are read by header name, not column position, so reordering
    or adding columns in the spreadsheet doesn't break parsing.
    """
    xlsx_path = ROOT / cfg["input"].get("schedule_xlsx_path", "data/schedule.xlsx")
    if not xlsx_path.exists():
        return []
    try:
        import openpyxl
    except ImportError:
        return []

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    def find_sheet(candidates: list[str]) -> str | None:
        normalized = {normalize_header(n): n for n in wb.sheetnames}
        for candidate in candidates:
            match = normalized.get(normalize_header(candidate))
            if match:
                return match
        return None

    presenters_by_session: dict[str, list[dict]] = {}
    presenter_sheet = find_sheet(_PRESENTER_SHEET_NAMES)
    if presenter_sheet:
        rows = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in wb[presenter_sheet].iter_rows(values_only=True)
        ]
        header_index = build_header_index(rows[0]) if rows else {}
        last_idx = find_header_indices(header_index, ["Last name", "Nachname"])
        first_idx = find_header_indices(header_index, ["First name", "Vorname"])
        topic_idx = find_header_indices(header_index, ["Topic", "Thema"])
        session_idx = find_header_indices(header_index, ["Session", "Modul"])
        for row in rows[1:]:
            session = get_cell(row, session_idx)
            if not session:
                continue
            name = normalize_case(
                " ".join(p for p in (get_cell(row, first_idx), get_cell(row, last_idx)) if p)
            )
            topic = normalize_case(get_cell(row, topic_idx))
            presenters_by_session.setdefault(session, []).append({"name": name, "topic": topic})

    schedule = []
    schedule_sheet = find_sheet(_SCHEDULE_SHEET_NAMES)
    if schedule_sheet:
        ws_s = wb[schedule_sheet]
        raw_rows = list(ws_s.iter_rows(values_only=True))
        header = [str(c) if c is not None else "" for c in raw_rows[0]] if raw_rows else []
        header_index = build_header_index(header)
        module_idx = find_header_indices(header_index, ["Module", "Modul"])
        time_idx = find_header_indices(header_index, ["Time", "Zeit"])
        if module_idx and time_idx:
            module_col, time_col = module_idx[0], time_idx[0]
            for row in raw_rows[1:]:
                if not row or module_col >= len(row):
                    continue
                module = str(row[module_col]).strip() if row[module_col] else ""
                if not module:
                    continue
                time_val = row[time_col] if time_col < len(row) else None
                if hasattr(time_val, "strftime"):
                    time_str = time_val.strftime("%H:%M")
                else:
                    time_str = str(time_val) if time_val else ""
                presenters = presenters_by_session.get(module, [])
                schedule.append({
                    "module": module,
                    "time": time_str,
                    "is_session": bool(presenters),
                    "presenters": presenters,
                })
    return schedule


def load_text_block(cfg: dict, docx_config_key: str, default_docx_name: str, html_prefix: str, lang_code: str) -> str:
    """Load a prose text block, preferring a mammoth-converted docx for DE.

    Content blocks (intro, location, and any future ones) all follow the same
    convention: a docx that's the DE source of truth (path from
    input.<docx_config_key>, e.g. input.intro_docx_path), auto-converted at
    build time, plus a hand-maintained data/<html_prefix>-<lang>.html for
    other languages, falling back to data/<html_prefix>.html if that's absent.
    """
    docx_path = ROOT / cfg["input"].get(docx_config_key, f"data/{default_docx_name}")
    if lang_code == "de" and docx_path.exists():
        try:
            import mammoth
            with open(docx_path, "rb") as fh:
                return mammoth.convert_to_html(fh).value
        except Exception:
            pass
    lang_path = ROOT / "data" / f"{html_prefix}-{lang_code}.html"
    fallback_path = ROOT / "data" / f"{html_prefix}.html"
    text_path = lang_path if lang_path.exists() else fallback_path
    return text_path.read_text(encoding="utf-8") if text_path.exists() else ""


def load_intro(cfg: dict, lang_code: str) -> str:
    return load_text_block(cfg, "intro_docx_path", "intro.docx", "intro", lang_code)


def load_location(cfg: dict, lang_code: str) -> str:
    return load_text_block(cfg, "location_docx_path", "location.docx", "location", lang_code)


def render(cfg: dict, sessions: list[dict], out_dir: Path) -> None:
    env = Environment(
        loader=FileSystemLoader(TEMPLATES),
        autoescape=select_autoescape(["html", "xml"]),
    )
    total = sum(len(s["entries"]) for s in sessions)
    out_dir.mkdir(parents=True, exist_ok=True)

    logo_src = ROOT / cfg["conference"].get("logo_path", "data/logo.png")
    has_logo = logo_src.exists()
    if has_logo:
        shutil.copy(logo_src, out_dir / "logo.png")

    try:
        from weasyprint import HTML as WPHtml
    except ImportError as exc:
        raise RuntimeError(
            "WeasyPrint is unavailable. Install the required system and Python dependencies "
            "to generate the PDF."
        ) from exc

    i18n = cfg.get("i18n") or {}
    generated = date.today().isoformat()
    schedule = load_schedule(cfg)

    for lang_code, t in i18n.items():
        intro_html = load_intro(cfg, lang_code)
        location_html = load_location(cfg, lang_code)

        html_name = "index.html" if lang_code == "de" else f"index-{lang_code}.html"
        ctx = {
            "conf": cfg["conference"],
            "sessions": sessions,
            "total": total,
            "generated": generated,
            "has_logo": has_logo,
            "intro": intro_html,
            "location": location_html,
            "schedule": schedule,
            "t": t,
            "lang": t.get("lang_attr", lang_code),
        }

        (out_dir / html_name).write_text(
            env.get_template("boc_web.html.j2").render(**ctx), encoding="utf-8"
        )

        pdf_name = t.get("pdf_file", f"book_of_contents_{lang_code}.pdf")
        WPHtml(
            string=env.get_template("boc_pdf.html.j2").render(**ctx),
            base_url=str(out_dir),
        ).write_pdf(str(out_dir / pdf_name))

    print(f"Rendered {total} entries → {out_dir}/  ({len(i18n)} language(s))")


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--input", "--csv", dest="input", help="override the input file path from config.yml")
    p.add_argument("--out", default="site", help="output directory (default: site)")
    args = p.parse_args()

    cfg = load_config()
    colmap = cfg["input"]["columns"]

    # Prefer xlsx over csv: try xlsx path first, fall back to csv_path from config.
    input_override = args.input
    if input_override:
        input_path = ROOT / input_override
        suffix = input_path.suffix.lower()
    else:
        xlsx_path = ROOT / cfg["input"].get("xlsx_path", "data/submissions.xlsx")
        csv_path  = ROOT / cfg["input"]["csv_path"]
        if xlsx_path.exists():
            input_path, suffix = xlsx_path, ".xlsx"
        else:
            input_path, suffix = csv_path, ".csv"

    if not input_path.exists():
        print(f"Input file not found: {input_path}", file=sys.stderr)
        return 1

    if suffix == ".xlsx":
        submissions = read_xlsx(input_path, colmap)
    else:
        submissions = read_csv(input_path, colmap)

    if not submissions:
        print("No submissions found in input file.", file=sys.stderr)
        return 1

    sessions = group_by_session(submissions, cfg)
    render(cfg, sessions, ROOT / args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
